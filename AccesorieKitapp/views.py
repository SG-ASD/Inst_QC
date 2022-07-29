from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.views.generic import UpdateView

from FinishedInspectionapp.models import FinishInspection
from Settingsapp.models import Settings
from .models import Accessories, AccessoriesFiles
from Inspectionapp.models import Inspection, Inspection_Category
from .forms import AccList1_UpdateForm, AccList2_UpdateForm, AccList3_UpdateForm
from openpyxl.styles import Alignment
import openpyxl
from QC_util import Util
import os
# Create your views here.


# 설명 : Seegene STARlet 악세서리 키트 첫번째 화면
# 작성자 : 이신후
# 날짜 : 2022/06/09
@method_decorator(login_required, 'get')
@method_decorator(login_required, 'post')
class AccKitUpdateView_first(UpdateView):
    model = Accessories
    form_class = AccList1_UpdateForm
    template_name = 'AccesorieKitapp/AccKitUpdateView_first.html'
    context_object_name = 'target_Acc_first'

    def get_object(self):
        object = get_object_or_404(Accessories, Instrument_SN=self.kwargs['Instrument_SN'])
        return object

    @transaction.atomic
    def get_context_data(self, **kwargs):
        context = super(AccKitUpdateView_first, self).get_context_data(**kwargs)
        context["inspection_category"] = Inspection_Category.objects.distinct().values_list('Category', flat=True)
        context["inspection_subcategory"] = Inspection_Category.objects.filter(Category="Electrical Test").values_list('Subcategory', flat=True)
        context['target_Acc_first_files'] = AccessoriesFiles.objects.filter(Instrument_SN=self.kwargs.get('Instrument_SN'))
        return context

    def get_success_url(self):
        return reverse("AccesorieKitapp:update_AccKit2", kwargs={"Instrument_SN": self.object.Instrument_SN_id})

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        instrument_SN = self.kwargs.get('Instrument_SN')
        Path = Settings.objects.get(pk=2)
        Path = Path.Path.replace('\\\\10.10.102.76\\장비품질관리팀\\품질관리_장비inspection\\','')
        if request.method == 'POST':
            form = AccList1_UpdateForm(request.POST)  # request된 폼


            if form.is_valid():  # 폼이 유효하면
                # db에 값 저장
                form_instance = get_object_or_404(Accessories, Instrument_SN=instrument_SN)  # 현재 Inspection 인스턴스를 불러온다.
                form_instance.Unpack_Instructions = request.POST.get('Unpack_Instructions')
                form_instance.Installation_Qualification = request.POST.get('Installation_Qualification')
                form_instance.Final_Configuation = request.POST.get('Final_Configuation')
                form_instance.Measurement_Protocol = request.POST.get('Unpack_Instructions')
                form_instance.EU_Declaration = request.POST.get('EU_Declaration')
                form_instance.DeclarationSTARlet = request.POST.get('DeclarationSTARlet')
                form_instance.Declaration_Quality = request.POST.get('Declaration_Quality')
                form_instance.Packing_Checklist = request.POST.get('Packing_Checklist')
                form_instance.Maintenance = request.POST.get('Maintenance')
                form_instance.Channel_Adjust = request.POST.get('Channel_Adjust')
                form_instance.Test_Report = request.POST.get('Test_Report')
                form_instance.Loading_Table = request.POST.get('Loading_Table')
                form_instance.Side_Cover_right = request.POST.get('Side_Cover_right')
                form_instance.Side_Cover_left = request.POST.get('Side_Cover_left')
                form_instance.Rear_Piexiglas = request.POST.get('Rear_Piexiglas')
                form_instance.Top_Plexiglas = request.POST.get('Top_Plexiglas')
                form_instance.Left_Plexiglas = request.POST.get('Left_Plexiglas')
                form_instance.Right_Plexiglas = request.POST.get('Right_Plexiglas')

                # 파일 upload
                # NAS_path = r"\home\windows\품질관리_장비inspection\01 검사 성적서 관리\2022 검사 성적서\QC SW 테스트"  # NAS 폴더 경로
                NAS_path = rf"\home\windows\품질관리_장비inspection\{Path}"  # NAS 폴더 경로
                path = NAS_path + '\\' + instrument_SN
                path = path.replace('\\', '/')
                file_instance = get_object_or_404(AccessoriesFiles, Instrument_SN=instrument_SN)  # 현재 Inspection 인스턴스를 불러온다.

                if request.FILES.getlist('Unpack_Instructions_Files'):
                    Unpack_Instructions_Files = request.FILES.getlist('Unpack_Instructions_Files')
                    Unpack_Instructions_Files_name = Util.upload_files(self, Unpack_Instructions_Files, path, '[Document] Unpacking Instructions, ML STAR', True)
                    file_instance.Unpack_Instructions_Files = Unpack_Instructions_Files_name
                else:
                    file_instance.Unpack_Instructions_Files = ""

                if request.FILES.getlist('Installation_Qualification_Files'):
                    Installation_Qualification_Files = request.FILES.getlist('Installation_Qualification_Files')
                    Installation_Qualification_Files_name = Util.upload_files(self, Installation_Qualification_Files, path, '[Document] Installation Qualification, ML STAR, IVD', True)
                    file_instance.Installation_Qualification_Files = Installation_Qualification_Files_name
                else:
                    file_instance.Installation_Qualification_Files = ""

                if request.FILES.getlist('Final_Configuation_Files'):
                    Final_Configuation_Files = request.FILES.getlist('Final_Configuation_Files')
                    Final_Configuation_Files_name = Util.upload_files(self, Final_Configuation_Files, path, '[Document] Final Configuration TEST STAR-Checklist', True)
                    file_instance.Final_Configuation_Files = Final_Configuation_Files_name
                else:
                    file_instance.Final_Configuation_Files = ""

                if request.FILES.getlist('Measurement_Protocol_Files'):
                    Measurement_Protocol_Files = request.FILES.getlist('Measurement_Protocol_Files')
                    Measurement_Protocol_Files_name = Util.upload_files(self, Measurement_Protocol_Files, path, '[Document] Measurement Protocol ML STAR IVD', True)
                    file_instance.Measurement_Protocol_Files = Measurement_Protocol_Files_name
                else:
                    file_instance.Measurement_Protocol_Files = ""

                if request.FILES.getlist('EU_Declaration_Files'):
                    EU_Declaration_Files = request.FILES.getlist('EU_Declaration_Files')
                    EU_Declaration_Files_name = Util.upload_files(self, EU_Declaration_Files, path, '[Document] EU Declaration of Conformity, ML STAR, IVD', True)
                    file_instance.EU_Declaration_Files = EU_Declaration_Files_name
                else:
                    file_instance.EU_Declaration_Files = ""

                if request.FILES.getlist('Declaration_Quality_Files'):
                    Declaration_Quality_Files = request.FILES.getlist('Declaration_Quality_Files')
                    Declaration_Quality_Files_name = Util.upload_files(self, Declaration_Quality_Files, path, '[Document] Declaration of Quality, Base Assembly Unit', True)
                    file_instance.Declaration_Quality_Files = Declaration_Quality_Files_name
                else:
                    file_instance.Declaration_Quality_Files = ""

                if request.FILES.getlist('DeclarationSTARlet_Files'):
                    DeclarationSTARlet_Files = request.FILES.getlist('DeclarationSTARlet_Files')
                    DeclarationSTARlet_Files_name = Util.upload_files(self, DeclarationSTARlet_Files, path, '[Document] Declaration of Quality, ML STAR, IVD', True)
                    file_instance.DeclarationSTARlet_Files = DeclarationSTARlet_Files_name
                else:
                    file_instance.DeclarationSTARlet_Files = ""

                if request.FILES.getlist('Packing_Checklist_Files'):
                    Packing_Checklist_Files = request.FILES.getlist('Packing_Checklist_Files')
                    Packing_Checklist_Files_name = Util.upload_files(self, Packing_Checklist_Files, path, '[Document] STARLINE Packing Checklist', True)
                    file_instance.Packing_Checklist_Files = Packing_Checklist_Files_name
                else:
                    file_instance.Packing_Checklist_Files = ""

                if request.FILES.getlist('Maintenance_Files'):
                    Maintenance_Files = request.FILES.getlist('Maintenance_Files')
                    Maintenance_Files_name = Util.upload_files(self, Maintenance_Files, path, '[Document] Maintenance (Barcode, Position, Weekly)', True)
                    file_instance.Maintenance_Files = Maintenance_Files_name
                else:
                    file_instance.Maintenance_Files = ""

                if request.FILES.getlist('Channel_Adjust_Files'):
                    Channel_Adjust_Files = request.FILES.getlist('Channel_Adjust_Files')
                    Channel_Adjust_Files_name = Util.upload_files(self, Channel_Adjust_Files, path, '[Document] Channel Adjust Value for Guldi Device', True)
                    file_instance.Channel_Adjust_Files = Channel_Adjust_Files_name
                else:
                    file_instance.Channel_Adjust_Files = ""

                if request.FILES.getlist('Test_Report_Files'):
                    Test_Report_Files = request.FILES.getlist('Test_Report_Files')
                    Test_Report_Files_name = Util.upload_files(self, Test_Report_Files, path, '[Document] STARLINE Electrical Safety Test Report', True)
                    file_instance.Test_Report_Files = Test_Report_Files_name
                else:
                    file_instance.Test_Report_Files = ""

                if request.FILES.getlist('Loading_Table_Files'):
                    Loading_Table_Files = request.FILES.getlist('Loading_Table_Files')
                    Loading_Table_Files_name = Util.upload_files(self, Loading_Table_Files, path, '173240_Loading Table, 30T', True)
                    file_instance.Loading_Table_Files = Loading_Table_Files_name
                else:
                    file_instance.Loading_Table_Files = ""

                if request.FILES.getlist('Side_Cover_right_Files'):
                    Side_Cover_right_Files = request.FILES.getlist('Side_Cover_right_Files')
                    Side_Cover_right_Files_name = Util.upload_files(self, Side_Cover_right_Files, path, '185091-RNO_Side Cover Plexiglas right(Flip up)', True)
                    file_instance.Side_Cover_right_Files = Side_Cover_right_Files_name
                else:
                    file_instance.Side_Cover_right_Files = ""

                if request.FILES.getlist('Side_Cover_left_Files'):
                    Side_Cover_left_Files = request.FILES.getlist('Side_Cover_left_Files')
                    Side_Cover_left_Files_name = Util.upload_files(self, Side_Cover_left_Files, path, '186571_Side Cover Plexiglas left', True)
                    file_instance.Side_Cover_left_Files = Side_Cover_left_Files_name
                else:
                    file_instance.Side_Cover_left_Files = ""

                if request.FILES.getlist('Rear_Piexiglas_Files'):
                    Rear_Piexiglas_Files = request.FILES.getlist('Rear_Piexiglas_Files')
                    Rear_Piexiglas_Files_name = Util.upload_files(self, Rear_Piexiglas_Files, path, '186572_REAR Plexiglas Cover 30T(for STARlet)', True)
                    file_instance.Rear_Piexiglas_Files = Rear_Piexiglas_Files_name
                else:
                    file_instance.Rear_Piexiglas_Files = ""

                if request.FILES.getlist('Top_Plexiglas_Files'):
                    Top_Plexiglas_Files = request.FILES.getlist('Top_Plexiglas_Files')
                    Top_Plexiglas_Files_name = Util.upload_files(self, Top_Plexiglas_Files, path, '186573_TOP Panel Plexiglas Cover 30T(for STARlet)', True)
                    file_instance.Top_Plexiglas_Files = Top_Plexiglas_Files_name
                else:
                    file_instance.Top_Plexiglas_Files = ""

                if request.FILES.getlist('Left_Plexiglas_Files'):
                    Left_Plexiglas_Files = request.FILES.getlist('Left_Plexiglas_Files')
                    Left_Plexiglas_Files_name = Util.upload_files(self, Left_Plexiglas_Files, path, '190201_LEFT REAR CORNER COVER (with Black)', True)
                    file_instance.Left_Plexiglas_Files = Left_Plexiglas_Files_name
                else:
                    file_instance.Left_Plexiglas_Files = ""

                if request.FILES.getlist('Right_Plexiglas_Files'):
                    Right_Plexiglas_Files = request.FILES.getlist('Right_Plexiglas_Files')
                    Right_Plexiglas_Files_name = Util.upload_files(self, Right_Plexiglas_Files, path, '190202_RIGHT REAR CORNER COVER (with Black)', True)
                    file_instance.Right_Plexiglas_Files = Right_Plexiglas_Files_name
                else:
                    file_instance.Right_Plexiglas_Files = ""

                form_instance.save()
                file_instance.save()

                return redirect('AccesorieKitapp:update_AccKit2', instrument_SN)
            else:
                return redirect('AccesorieKitapp:update_AccKit1', instrument_SN)


# 설명 : Seegene STARlet 악세서리 키트 두번째 화면
# 작성자 : 이신후
# 날짜 : 2022/06/09
@method_decorator(login_required, 'get')
@method_decorator(login_required, 'post')
class AccKitUpdateView_second(UpdateView):
    model = Accessories
    form_class = AccList2_UpdateForm
    template_name = 'AccesorieKitapp/AccKitUpdateView_second.html'
    context_object_name = 'target_Acc_second'

    def get_object(self):
        object = get_object_or_404(Accessories, Instrument_SN=self.kwargs['Instrument_SN'])
        return object

    @transaction.atomic
    def get_context_data(self, **kwargs):
        context = super(AccKitUpdateView_second, self).get_context_data(**kwargs)
        context["inspection_category"] = Inspection_Category.objects.distinct().values_list('Category', flat=True)
        context["inspection_subcategory"] = Inspection_Category.objects.filter(Category="Electrical Test").values_list('Subcategory', flat=True)
        context['target_Acc_second_files'] = AccessoriesFiles.objects.filter(Instrument_SN=self.kwargs.get('Instrument_SN'))
        return context

    def get_success_url(self):
        return reverse("AccesorieKitapp:update_AccKit3", kwargs={"Instrument_SN": self.object.Instrument_SN_id})

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        instrument_SN = self.kwargs.get('Instrument_SN')
        Path = Settings.objects.get(pk=2)
        Path = Path.Path.replace('\\\\10.10.102.76\\장비품질관리팀\\품질관리_장비inspection\\','')
        if request.method == 'POST':
            form = AccList2_UpdateForm(request.POST)  # request된 폼

            if form.is_valid():  # 폼이 유효하면
                # db에 값 저장
                form_instance = get_object_or_404(Accessories, Instrument_SN=instrument_SN)  # 현재 Inspection 인스턴스를 불러온다.
                form_instance.Sensor_4L = request.POST.get('Sensor_4L')
                form_instance.Solid_Waste = request.POST.get('Solid_Waste')
                form_instance.USB_Liquid = request.POST.get('USB_Liquid')
                form_instance.USB_Solid_Liquid = request.POST.get('USB_Solid_Liquid')
                form_instance.USB_Venus = request.POST.get('USB_Venus')
                form_instance.Eppendorf = request.POST.get('Eppendorf')
                form_instance.Core_Gripper = request.POST.get('Core_Gripper')
                form_instance.Needle_Lot = request.POST.get('Needle_Lot')
                form_instance.Needle_Check = request.POST.get('Needle_Check')
                form_instance.Separator = request.POST.get('Separator')
                form_instance.Tip_Eject = request.POST.get('Tip_Eject')
                form_instance.Power_Cord7 = request.POST.get('Power_Cord7')
                form_instance.Power_CordP = request.POST.get('Power_CordP')
                form_instance.Cable_USB = request.POST.get('Cable_USB')
                form_instance.Fuse = request.POST.get('Fuse')
                form_instance.Screw = request.POST.get('Screw')
                form_instance.Fitting = request.POST.get('Fitting')
                form_instance.Liq_Waste = request.POST.get('Liq_Waste')

                # 파일 upload
                # NAS_path = r"\home\windows\품질관리_장비inspection\01 검사 성적서 관리\2022 검사 성적서\QC SW 테스트"  # NAS 폴더 경로
                NAS_path = rf"\home\windows\품질관리_장비inspection\{Path}"  # NAS 폴더 경로
                path = NAS_path + '\\' + instrument_SN
                path = path.replace('\\', '/')
                file_instance = get_object_or_404(AccessoriesFiles, Instrument_SN=instrument_SN)  # 현재 Inspection 인스턴스를 불러온다.

                if request.FILES.getlist('Sensor_4L_Files'):
                    Sensor_4L_Files = request.FILES.getlist('Sensor_4L_Files')
                    Sensor_4L_Files_name = Util.upload_files(self, Sensor_4L_Files, path, '96718-01_ASSY, LIQ WASTE, 5L, 4L-SENSOR', True)
                    file_instance.Sensor_4L_Files = Sensor_4L_Files_name
                else:
                    file_instance.Sensor_4L_Files = ""

                if request.FILES.getlist('Solid_Waste_Files'):
                    Solid_Waste_Files = request.FILES.getlist('Solid_Waste_Files')
                    Solid_Waste_Files_name = Util.upload_files(self, Solid_Waste_Files, path, '53873-07_SOLID WASTE CONTAINER', True)
                    file_instance.Solid_Waste_Files = Solid_Waste_Files_name
                else:
                    file_instance.Solid_Waste_Files = ""

                if request.FILES.getlist('USB_Liquid_Files'):
                    USB_Liquid_Files = request.FILES.getlist('USB_Liquid_Files')
                    USB_Liquid_name_Files = Util.upload_files(self, USB_Liquid_Files, path, '54698-01_[USB]Liquid and Tip Waste System, User Guide, Rev B', True)
                    file_instance.USB_Liquid_Files = USB_Liquid_name_Files
                else:
                    file_instance.USB_Liquid_Files = ""

                if request.FILES.getlist('USB_Solid_Liquid_Files'):
                    USB_Solid_Liquid_Files = request.FILES.getlist('USB_Solid_Liquid_Files')
                    USB_Solid_Liquid_Files_name = Util.upload_files(self, USB_Solid_Liquid_Files, path, '54802-01_[CD] Solid-Liquid Waste Sensor V1.3, Rev B', True)
                    file_instance.USB_Solid_Liquid_Files = USB_Solid_Liquid_Files_name
                else:
                    file_instance.USB_Solid_Liquid_Files = ""

                if request.FILES.getlist('USB_Venus_Files'):
                    USB_Venus_Files = request.FILES.getlist('USB_Venus_Files')
                    USB_Venus_Files_name = Util.upload_files(self, USB_Venus_Files, path, '911265-01_[USB] VENUS FOUR CD SW & MANUAL STAR IVD 4.5, Rev X1', True)
                    file_instance.USB_Venus_Files = USB_Venus_Files_name
                else:
                    file_instance.USB_Venus_Files = ""

                if request.FILES.getlist('Eppendorf_Files'):
                    Eppendorf_Files = request.FILES.getlist('Eppendorf_Files')
                    Eppendorf_Files_name = Util.upload_files(self, Eppendorf_Files, path, '182238_32 Inserts for 1.5ml Eppendorf Cups', True)
                    file_instance.Eppendorf_Files = Eppendorf_Files_name
                else:
                    file_instance.Eppendorf_Files = ""

                if request.FILES.getlist('Core_Gripper_Files'):
                    Core_Gripper_Files = request.FILES.getlist('Core_Gripper_Files')
                    Core_Gripper_Files_name = Util.upload_files(self, Core_Gripper_Files, path, '184089_CORE Gripper (Screw x 4ea included)', True)
                    file_instance.Core_Gripper_Files = Core_Gripper_Files_name
                else:
                    file_instance.Core_Gripper_Files = ""

                if request.FILES.getlist('Needle_Check_Files'):
                    Needle_Check_Files = request.FILES.getlist('Needle_Check_Files')
                    Needle_Check_Files_name = Util.upload_files(self, Needle_Check_Files, path, '182136_Teaching Needle(8ea)', True)
                    file_instance.Needle_Check_Files = Needle_Check_Files_name
                else:
                    file_instance.Needle_Check_Files = ""

                if request.FILES.getlist('Separator_Files'):
                    Separator_Files = request.FILES.getlist('Separator_Files')
                    Separator_Files_name = Util.upload_files(self, Separator_Files, path, '53894-01_SEPARATOR, MAGNETIC, NUCLEO', True)
                    file_instance.Separator_Files = Separator_Files_name
                else:
                    file_instance.Separator_Files = ""

                if request.FILES.getlist('Tip_Eject_Files'):
                    Tip_Eject_Files = request.FILES.getlist('Tip_Eject_Files')
                    Tip_Eject_Files_name = Util.upload_files(self, Tip_Eject_Files, path, '58938-01_Tip Eject Plate(DEFLECTOR, TIP, LIQUID WASTE)', True)
                    file_instance.Tip_Eject_Files = Tip_Eject_Files_name
                else:
                    file_instance.Tip_Eject_Files = ""

                if request.FILES.getlist('Power_Cord7_Files'):
                    Power_Cord7_Files = request.FILES.getlist('Power_Cord7_Files')
                    Power_Cord7_Files_name = Util.upload_files(self, Power_Cord7_Files, path, '3892-01_POWER CORD, STD, CEE', True)
                    file_instance.Power_Cord7_Files = Power_Cord7_Files_name
                else:
                    file_instance.Power_Cord7_Files = ""

                if request.FILES.getlist('Power_CordP_Files'):
                    Power_CordP_Files = request.FILES.getlist('Power_CordP_Files')
                    Power_CordP_Files_name = Util.upload_files(self, Power_CordP_Files, path, '3892-05_POWER CORD, USA, NEMA 5-15P', True)
                    file_instance.Power_CordP_Files = Power_CordP_Files_name
                else:
                    file_instance.Power_CordP_Files = ""

                if request.FILES.getlist('Cable_USB_Files'):
                    Cable_USB_Files = request.FILES.getlist('Cable_USB_Files')
                    Cable_USB_Files_name = Util.upload_files(self, Cable_USB_Files, path, '355171_CABLE USB 2.0 A-B 3M FERRIT', True)
                    file_instance.Cable_USB_Files = Cable_USB_Files_name
                else:
                    file_instance.Cable_USB_Files = ""

                if request.FILES.getlist('Fuse_Files'):
                    Fuse_Files = request.FILES.getlist('Fuse_Files')
                    Fuse_Files_name = Util.upload_files(self, Fuse_Files, path, '363013_FUSE 6.3AT', True)
                    file_instance.Fuse_Files = Fuse_Files_name
                else:
                    file_instance.Fuse_Files = ""

                if request.FILES.getlist('Screw_Files'):
                    Screw_Files = request.FILES.getlist('Screw_Files')
                    Screw_Files_name = Util.upload_files(self, Screw_Files, path, '420629_SCREW M4x16', True)
                    file_instance.Power_Cord7_Files = Screw_Files_name
                else:
                    file_instance.Power_Cord7_Files = ""

                if request.FILES.getlist('Fitting_Files'):
                    Fitting_Files = request.FILES.getlist('Fitting_Files')
                    Fitting_Files_name = Util.upload_files(self, Fitting_Files, path, '53866-01_FITTING, QIK-CON, IDX', True)
                    file_instance.Fitting_Files = Fitting_Files_name
                else:
                    file_instance.Fitting_Files = ""

                if request.FILES.getlist('Liq_Waste_Files'):
                    Liq_Waste_Files = request.FILES.getlist('Liq_Waste_Files')
                    Liq_Waste_Files_name = Util.upload_files(self, Liq_Waste_Files, path, '54200-01_ASSY, FILTER, AEROSOL, LIQ WASTE', True)
                    file_instance.Liq_Waste_Files = Liq_Waste_Files_name
                else:
                    file_instance.Liq_Waste_Files = ""

                form_instance.save()
                file_instance.save()

                return redirect('AccesorieKitapp:update_AccKit3', instrument_SN)
            else:
                return redirect('AccesorieKitapp:update_AccKit2', instrument_SN)


# 설명 : Seegene STARlet 악세서리 키트 세번째 화면
# 작성자 : 이신후
# 날짜 : 2022/06/09
@method_decorator(login_required, 'get')
@method_decorator(login_required, 'post')
class AccKitUpdateView_third(UpdateView):
    model = Accessories
    model_Inspection = Inspection
    form_class = AccList3_UpdateForm
    template_name = 'AccesorieKitapp/AccKitUpdateView_third.html'
    context_object_name = 'target_Acc_third'

    def get_object(self):
        HHS_SN = get_object_or_404(Inspection, Instrument_SN=self.kwargs['Instrument_SN']).Perfomance_HHS_SN
        Accessories.objects.filter(Instrument_SN=self.kwargs['Instrument_SN']).update(HHS_SN=HHS_SN)
        object = get_object_or_404(Accessories, Instrument_SN=self.kwargs['Instrument_SN'])
        return object

    @transaction.atomic
    def get_context_data(self, **kwargs):
        context = super(AccKitUpdateView_third, self).get_context_data(**kwargs)
        context["inspection"] = Inspection.objects.filter(Instrument_SN=self.kwargs['Instrument_SN'])
        context["inspection_category"] = Inspection_Category.objects.distinct().values_list('Category', flat=True)
        context["inspection_subcategory"] = Inspection_Category.objects.filter(Category="Electrical Test").values_list('Subcategory', flat=True)
        context['target_Acc_third_files'] = AccessoriesFiles.objects.filter(Instrument_SN=self.kwargs.get('Instrument_SN'))
        return context

    def get_success_url(self):
        object_Inspection = get_object_or_404(Inspection, Instrument_SN=self.kwargs['Instrument_SN'])
        CompleteDt = self.request.POST.get("CompleteDt")
        # Inspection.objects.filter(Instrument_SN=object_Inspection.Instrument_SN_id).update(CompleteDt=CompleteDt)
        # 악세서리 키트에서 완제품 넘어갈 경우, 자동 Excel file 데이터 입력 기능
        # self.Excel_Inspection1()
        # self.Excel_Inspection2()
        # self.Excel_Inspection3()
        # self.Excel_Inspection4()
        return reverse("FinishedInspectionapp:update_finish1", kwargs={"Instrument_SN": self.object.Instrument_SN_id})

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        instrument_SN = self.kwargs.get('Instrument_SN')
        Path = Settings.objects.get(pk=2)
        Path = Path.Path.replace('\\\\10.10.102.76\\장비품질관리팀\\품질관리_장비inspection\\','')

        if request.method == 'POST':
            form = AccList3_UpdateForm(request.POST)  # request된 폼

            if form.is_valid():  # 폼이 유효하면
                # db에 값 저장
                form_instance = get_object_or_404(Accessories, Instrument_SN=instrument_SN)  # 현재 악세서리 인스턴스를 불러온다.
                form_inspection_instance = get_object_or_404(Inspection, Instrument_SN=instrument_SN)  # 현재 Inspection 인스턴스를 불러온다.
                form_finish_instance = get_object_or_404(FinishInspection, Instrument_SN=instrument_SN)  # 현재 Inspection 인스턴스를 불러온다.

                form_inspection_instance.CompleteDt = request.POST.get('CompleteDt')

                form_finish_instance.Start_Date = request.POST.get('Start_Date')

                form_instance.XArm = request.POST.get('XArm')
                form_instance.Bar_Left = request.POST.get('Bar_Left')
                form_instance.Bar_Top = request.POST.get('Bar_Top')
                form_instance.Bag = request.POST.get('Bag')
                form_instance.Tube_32 = request.POST.get('Tube_32')
                form_instance.TipRack_5 = request.POST.get('TipRack_5')
                form_instance.MFX4 = request.POST.get('MFX4')
                form_instance.SEEG = request.POST.get('SEEG')
                form_instance.FLHX = request.POST.get('FLHX')
                form_instance.SOHX = request.POST.get('SOHX')
                form_instance.HHS_Plate = request.POST.get('HHS_Plate')
                form_instance.MTP = request.POST.get('MTP')
                form_instance.Verify_Kit = request.POST.get('Verify_Kit')
                form_instance.STD = request.POST.get('STD')
                form_instance.High = request.POST.get('High')
                form_instance.HHS_Check = request.POST.get('HHS_Check')
                form_instance.HHS_Manual = request.POST.get('HHS_Manual')
                form_instance.Stop_Barcode = request.POST.get('Stop_Barcode')
                form_instance.Remark = request.POST.get('Remark')

                # 파일 upload
                # NAS_path = r"\home\windows\품질관리_장비inspection\01 검사 성적서 관리\2022 검사 성적서\QC SW 테스트"  # NAS 폴더 경로
                NAS_path = rf"\home\windows\품질관리_장비inspection\{Path}"  # NAS 폴더 경로
                path = NAS_path + '\\' + instrument_SN
                path = path.replace('\\', '/')
                file_instance = get_object_or_404(AccessoriesFiles, Instrument_SN=instrument_SN)  # 현재 Inspection 인스턴스를 불러온다.

                if request.FILES.getlist('XArm_Files'):
                    XArm_Files = request.FILES.getlist('XArm_Files')
                    XArm_Files_name = Util.upload_files(self, XArm_Files, path, '(Channel support)X-Arm Holding brakets', True)
                    file_instance.XArm_Files = XArm_Files_name
                else:
                    file_instance.XArm_Files = ""

                if request.FILES.getlist('Bar_Left_Files'):
                    Bar_Left_Files = request.FILES.getlist('Bar_Left_Files')
                    Bar_Left_Files_name = Util.upload_files(self, Bar_Left_Files, path, '190148_TRANSPORT SUPPORT BAR LEFT', True)
                    file_instance.Bar_Left_Files = Bar_Left_Files_name
                else:
                    file_instance.Bar_Left_Files = ""

                if request.FILES.getlist('Bar_Top_Files'):
                    Bar_Top_Files = request.FILES.getlist('Bar_Top_Files')
                    Bar_Top_Files_name = Util.upload_files(self, Bar_Top_Files, path, '186675_Support Bar on Top (for Top Plexiglas Cover)', True)
                    file_instance.Bar_Top_Files = Bar_Top_Files_name
                else:
                    file_instance.Bar_Top_Files = ""

                if request.FILES.getlist('Bag_Files'):
                    Bag_Files = request.FILES.getlist('Bag_Files')
                    Bag_Files_name = Util.upload_files(self, Bag_Files, path, '53686-01_BAG, BIO-HAZARD, 14.5 x 8.2 x 18', True)
                    file_instance.Bag_Files = Bag_Files_name
                else:
                    file_instance.Bag_Files = ""

                if request.FILES.getlist('Tube_32_Files'):
                    Tube_32_Files = request.FILES.getlist('Tube_32_Files')
                    Tube_32_Files_name = Util.upload_files(self, Tube_32_Files, path, '173410_3 Carrier for 32 Tubes(11x60 - 14x120mm)', True)
                    file_instance.Tube_32_Files = Tube_32_Files_name
                else:
                    file_instance.Tube_32_Files = ""

                if request.FILES.getlist('TipRack_5_Files'):
                    TipRack_5_Files = request.FILES.getlist('TipRack_5_Files')
                    TipRack_5_Files_name = Util.upload_files(self, TipRack_5_Files, path, '182085_TIP CARRIER, 5 TIP RACKS', True)
                    file_instance.TipRack_5_Files = TipRack_5_Files_name
                else:
                    file_instance.TipRack_5_Files = ""

                if request.FILES.getlist('MFX4_Files'):
                    MFX4_Files = request.FILES.getlist('MFX4_Files')
                    MFX4_Files_name = Util.upload_files(self, MFX4_Files, path, '97076-01_ASSY, CAR, 7T, HHS, 4MFX', True)
                    file_instance.MFX4_Files = MFX4_Files_name
                else:
                    file_instance.MFX4_Files = ""

                if request.FILES.getlist('SEEG_Files'):
                    SEEG_Files = request.FILES.getlist('SEEG_Files')
                    SEEG_Files_name = Util.upload_files(self, SEEG_Files, path, '97432-01_PED, ADAPTER, SMALL TUBE, SEEG', True)
                    file_instance.SEEG_Files = SEEG_Files_name
                else:
                    file_instance.SEEG_Files = ""

                if request.FILES.getlist('FLHX_Files'):
                    FLHX_Files = request.FILES.getlist('FLHX_Files')
                    FLHX_Files_name = Util.upload_files(self, FLHX_Files, path, '403453_SCREW, FLHX M3X8, DIN7991, A2SS', True)
                    file_instance.FLHX_Files = FLHX_Files_name
                else:
                    file_instance.FLHX_Files = ""

                if request.FILES.getlist('SOHX_Files'):
                    SOHX_Files = request.FILES.getlist('SOHX_Files')
                    SOHX_Files_name = Util.upload_files(self, SOHX_Files, path, '400263_SCREW, SOHX, M3X8, DIN912, A2SS', True)
                    file_instance.SOHX_Files = SOHX_Files_name
                else:
                    file_instance.SOHX_Files = ""

                if request.FILES.getlist('HHS_Plate_Files'):
                    HHS_Plate_Files = request.FILES.getlist('HHS_Plate_Files')
                    HHS_Plate_Files_name = Util.upload_files(self, HHS_Plate_Files, path, 'HHS Base Plate', True)
                    file_instance.HHS_Plate_Files = HHS_Plate_Files_name
                else:
                    file_instance.HHS_Plate_Files = ""

                if request.FILES.getlist('MTP_Files'):
                    MTP_Files = request.FILES.getlist('MTP_Files')
                    MTP_Files_name = Util.upload_files(self, MTP_Files, path, '97319-01_ASSY, CAR, MFX, 6T, REAG, DW, MTP', True)
                    file_instance.MTP_Files = MTP_Files_name
                else:
                    file_instance.MTP_Files = ""

                if request.FILES.getlist('Verify_Kit_Files'):
                    Verify_Kit_Files = request.FILES.getlist('Verify_Kit_Files')
                    Verify_Kit_Files_name = Util.upload_files(self, Verify_Kit_Files, path, '62964-03_RENO KIT 3, STAR VERIFY KIT', True)
                    file_instance.Verify_Kit_Files = Verify_Kit_Files_name
                else:
                    file_instance.Verify_Kit_Files = ""

                if request.FILES.getlist('STD_Files'):
                    STD_Files = request.FILES.getlist('STD_Files')
                    STD_Files_name = Util.upload_files(self, STD_Files, path, '235969_STD VOL CO-RE TIPS FIL 5x96', True)
                    file_instance.STD_Files = STD_Files_name
                else:
                    file_instance.STD_Files = ""

                if request.FILES.getlist('High_Files'):
                    High_Files = request.FILES.getlist('High_Files')
                    High_Files_name = Util.upload_files(self, High_Files, path, '235970_HIGH VOL EO-RE TIPS FIL 5x96', True)
                    file_instance.High_Files = High_Files_name
                else:
                    file_instance.High_Files = ""

                if request.FILES.getlist('HHS_Check_Files'):
                    HHS_Check_Files = request.FILES.getlist('HHS_Check_Files')
                    HHS_Check_Files_name = Util.upload_files(self, HHS_Check_Files, path, '199034_Hamiltion Heater & Shaker', True)
                    file_instance.HHS_Check_Files = HHS_Check_Files_name
                else:
                    file_instance.HHS_Check_Files = ""

                if request.FILES.getlist('HHS_Manual_Files'):
                    HHS_Manual_Files = request.FILES.getlist('HHS_Manual_Files')
                    HHS_Manual_Files_name = Util.upload_files(self, HHS_Manual_Files, path, '99088-24_[USB] Hamilton Heater Shaker Manual', True)
                    file_instance.HHS_Manual_Files = HHS_Manual_Files_name
                else:
                    file_instance.HHS_Manual_Files = ""

                if request.FILES.getlist('Stop_Barcode_Files'):
                    Stop_Barcode_Files = request.FILES.getlist('Stop_Barcode_Files')
                    Stop_Barcode_Files_name = Util.upload_files(self, Stop_Barcode_Files, path, 'Stop Barcode', True)
                    file_instance.Stop_Barcode_Files = Stop_Barcode_Files_name
                else:
                    file_instance.Stop_Barcode_Files = ""

                if request.FILES.getlist('Remark_Files'):
                    Remark_Files = request.FILES.getlist('Remark_Files')
                    Remark_Files_name = Util.upload_files(self, Remark_Files, path, 'Remark_Files', True)
                    file_instance.Remark_Files = Remark_Files_name
                else:
                    file_instance.Remark_Files = ""

                form_inspection_instance.save()
                form_finish_instance.save()
                form_instance.save()
                file_instance.save()

                return redirect('FinishedInspectionapp:update_finish1', instrument_SN)
            else:
                return redirect('AccesorieKitapp:update_AccKit3', instrument_SN)

    def Excel_Inspection1(self):
        object_Inspection = get_object_or_404(Inspection, Instrument_SN=self.kwargs['Instrument_SN'])
        # 수입검사 성적서 엑셀 파일에 데이터 삽입
        wb = openpyxl.load_workbook('D:\Inst_QC\Report.xlsx', data_only=True)
        sheet = wb.active
        sheet = wb['1']
        sheet['E10'] = object_Inspection.Inspector
        sheet['E11'] = object_Inspection.Start_Date
        sheet['E12'] = object_Inspection.CompleteDt
        sheet['E14'] = object_Inspection.SW_Version
        sheet['E15'] = object_Inspection.FV2_Version
        sheet['E16'] = object_Inspection.FW_PipetteCh
        sheet['E17'] = object_Inspection.FW_Xdrive
        sheet['E18'] = object_Inspection.FW_Master

        sheet['K7'] = object_Inspection.Name
        sheet['K8'] = object_Inspection.Computer_SN

        # Inspection Appearance 부분
        if object_Inspection.Appearance_Shock_Watch == "No":
            sheet['G23'] = "■ No          □  Yes "
        elif object_Inspection.Appearance_Shock_Watch == "Yes":
            sheet['G23'] = "□ No          ■  Yes "

        if object_Inspection.Appearance_Binding == "No":
            sheet['G24'] = "■ No          □  Yes "
        elif object_Inspection.Appearance_Binding == "Yes":
            sheet['G24'] = "□ No          ■  Yes "

        if object_Inspection.Appearance_Labels == "No":
            sheet['G25'] = "■ No          □  Yes "
        elif object_Inspection.Appearance_Labels == "Yes":
            sheet['G25'] = "□ No          ■  Yes "

        if object_Inspection.Appearance_Packaging_Box == "No":
            sheet['G26'] = "■ No          □  Yes "
        elif object_Inspection.Appearance_Packaging_Box == "Yes":
            sheet['G26'] = "□ No          ■  Yes "

        if object_Inspection.Appearance_Wooden_Pallet == "No":
            sheet['G27'] = "■ No          □  Yes "
        elif object_Inspection.Appearance_Wooden_Pallet == "Yes":
            sheet['G27'] = "□ No          ■  Yes "

        if object_Inspection.Appearance_Transport_Jig == "No":
            sheet['G28'] = "■ No          □  Yes "
        elif object_Inspection.Appearance_Transport_Jig == "Yes":
            sheet['G28'] = "□ No          ■  Yes "

        # Appearance Front / Top / Right / Left /Back
        if object_Inspection.Appearance_Front == "Pass":
            sheet['K33'] = "■ Pass"
            sheet['K34'] = "□  Fail"
        elif object_Inspection.Appearance_Front == "Fail":
            sheet['K33'] = "□  Pass"
            sheet['K34'] = "■  Fail"

        if object_Inspection.Appearance_Top == "Pass":
            sheet['K36'] = "■ Pass"
            sheet['K37'] = "□  Fail"
        elif object_Inspection.Appearance_Top == "Fail":
            sheet['K36'] = "□  Pass"
            sheet['K37'] = "■  Fail"

        if object_Inspection.Appearance_Right == "Pass":
            sheet['K39'] = "■ Pass"
            sheet['K40'] = "□  Fail"
        elif object_Inspection.Appearance_Right == "Fail":
            sheet['K39'] = "□  Pass"
            sheet['K40'] = "■  Fail"

        if object_Inspection.Appearance_Left == "Pass":
            sheet['K42'] = "■ Pass"
            sheet['K43'] = "□  Fail"
        elif object_Inspection.Appearance_Left == "Fail":
            sheet['K42'] = "□  Pass"
            sheet['K43'] = "■  Fail"

        if object_Inspection.Appearance_Back == "Pass":
            sheet['K45'] = "■ Pass"
            sheet['K46'] = "□  Fail"
        elif object_Inspection.Appearance_Back == "Fail":
            sheet['K45'] = "□  Pass"
            sheet['K46'] = "■  Fail"

        wb.save('D:\Inst_QC\Report.xlsx')

    def Excel_Inspection2(self):
        # 성적서 2페이지
        object_Inspection = get_object_or_404(Inspection, Instrument_SN=self.kwargs['Instrument_SN'])
        # 수입검사 성적서 엑셀 파일에 데이터 삽입
        wb = openpyxl.load_workbook('D:\Inst_QC\Report.xlsx', data_only=True)
        sheet = wb.active
        sheet = wb['2']
        if object_Inspection.Appearance_Acc_Damage == "Yes":
            sheet['H3'] = "□ No          ■  Yes "
        elif object_Inspection.Appearance_Acc_Damage == "No":
            sheet['H3'] = "■ No          □  Yes "

        if object_Inspection.Appearance_Acc_Missing == "Yes":
            sheet['H4'] = "□ No          ■  Yes "
        elif object_Inspection.Appearance_Acc_Missing == "No":
            sheet['H4'] = "■ No          □  Yes "

        if object_Inspection.ElectricalTest_HiPotential == "Pass":
            sheet['H8'] = "■ Pass          □  Fail"
        elif object_Inspection.ElectricalTest_HiPotential == "Fail":
            sheet['H8'] = "□ Pass          ■  Fail"

        if object_Inspection.ElectricalTest_GroundResistance == "Pass":
            sheet['H9'] = "■ Pass          □  Fail"
        elif object_Inspection.ElectricalTest_GroundResistance == "Fail":
            sheet['H9'] = "□ Pass          ■  Fail"

        if object_Inspection.ElectricalTest_PowerSWFunction == "Pass":
            sheet['H10'] = "■ Pass          □  Fail"
        elif object_Inspection.ElectricalTest_PowerSWFunction == "Fail":
            sheet['H10'] = "□ Pass          ■  Fail"

        if object_Inspection.ElectricalTest_PowerLED == "Pass":
            sheet['H11'] = "■ Pass          □  Fail"
        elif object_Inspection.ElectricalTest_PowerLED == "Fail":
            sheet['H11'] = "□ Pass          ■  Fail"

        sheet['B14'] = "Used Channel Calibration Tool Management No : " + str(
            object_Inspection.HardWare_CalibrationTool)
        sheet['B15'] = "Used Barcode Carrier Management No : " + str(object_Inspection.HardWare_BarcodeCarrier)

        sheet[
            'H17'] = f"     Left :      {object_Inspection.HardWare_XArm_left}               Right :     {object_Inspection.HardWare_XArm_right}       "

        sheet['H18'] = object_Inspection.HardWare_XArm_Diff

        sheet['I20'] = f"X :     {object_Inspection.HardWare_Xdev1}   //  {object_Inspection.HardWare_Xtilt1}"
        sheet['I21'] = f"X :     {object_Inspection.HardWare_Xdev2}   //  {object_Inspection.HardWare_Xtilt2}"
        sheet['I22'] = f"X :     {object_Inspection.HardWare_Xdev3}   //  {object_Inspection.HardWare_Xtilt3}"
        sheet['I23'] = f"X :     {object_Inspection.HardWare_Xdev4}   //  {object_Inspection.HardWare_Xtilt4}"
        sheet['I24'] = f"X :     {object_Inspection.HardWare_Xdev5}   //  {object_Inspection.HardWare_Xtilt5}"
        sheet['I25'] = f"X :     {object_Inspection.HardWare_Xdev6}   //  {object_Inspection.HardWare_Xtilt6}"
        sheet['I26'] = f"X :     {object_Inspection.HardWare_Xdev7}   //  {object_Inspection.HardWare_Xtilt7}"
        sheet['I27'] = f"X :     {object_Inspection.HardWare_Xdev8}   //  {object_Inspection.HardWare_Xtilt8}"

        sheet['J20'] = f"Y :   {object_Inspection.HardWare_Ytilt1}"
        sheet['J21'] = f"Y :   {object_Inspection.HardWare_Ytilt2}"
        sheet['J22'] = f"Y :   {object_Inspection.HardWare_Ytilt3}"
        sheet['J23'] = f"Y :   {object_Inspection.HardWare_Ytilt4}"
        sheet['J24'] = f"Y :   {object_Inspection.HardWare_Ytilt5}"
        sheet['J25'] = f"Y :   {object_Inspection.HardWare_Ytilt6}"
        sheet['J26'] = f"Y :   {object_Inspection.HardWare_Ytilt7}"
        sheet['J27'] = f"Y :   {object_Inspection.HardWare_Ytilt8}"

        sheet['K20'] = object_Inspection.HardWare_SN1
        sheet['K21'] = object_Inspection.HardWare_SN2
        sheet['K22'] = object_Inspection.HardWare_SN3
        sheet['K23'] = object_Inspection.HardWare_SN4
        sheet['K24'] = object_Inspection.HardWare_SN5
        sheet['K25'] = object_Inspection.HardWare_SN6
        sheet['K26'] = object_Inspection.HardWare_SN7
        sheet['K27'] = object_Inspection.HardWare_SN8

        if object_Inspection.HardWare_Adjust == "Pass":
            sheet['E28'] = "■ Pass          □  Fail"
        elif object_Inspection.HardWare_Adjust == "Fail":
            sheet['E28'] = "□ Pass          ■  Fail"

        sheet[
            'I29'] = f"X :    {object_Inspection.HardWare_PIP_X1}                                Y : {object_Inspection.HardWare_PIP_Y1} "
        sheet[
            'I31'] = f"X :    {object_Inspection.HardWare_PIP_X2}                                Y : {object_Inspection.HardWare_PIP_Y2} "
        sheet[
            'I33'] = f"X :    {object_Inspection.HardWare_PIP_X3}                                Y : {object_Inspection.HardWare_PIP_Y3} "
        sheet[
            'I35'] = f"X :    {object_Inspection.HardWare_PIP_X4}                                Y : {object_Inspection.HardWare_PIP_Y4} "
        sheet[
            'I37'] = f"X :    {object_Inspection.HardWare_PIP_X5}                                Y : {object_Inspection.HardWare_PIP_Y5} "
        sheet[
            'I39'] = f"X :    {object_Inspection.HardWare_PIP_X6}                                Y : {object_Inspection.HardWare_PIP_Y6} "
        sheet[
            'I41'] = f"X :    {object_Inspection.HardWare_PIP_X7}                                Y : {object_Inspection.HardWare_PIP_Y7} "
        sheet[
            'I43'] = f"X :    {object_Inspection.HardWare_PIP_X8}                                Y : {object_Inspection.HardWare_PIP_Y8} "

        sheet['I30'] = f"Z : {object_Inspection.HardWare_PIP_Z1}"
        sheet['I32'] = f"Z : {object_Inspection.HardWare_PIP_Z2}"
        sheet['I34'] = f"Z : {object_Inspection.HardWare_PIP_Z3}"
        sheet['I36'] = f"Z : {object_Inspection.HardWare_PIP_Z4}"
        sheet['I38'] = f"Z : {object_Inspection.HardWare_PIP_Z5}"
        sheet['I40'] = f"Z : {object_Inspection.HardWare_PIP_Z6}"
        sheet['I42'] = f"Z : {object_Inspection.HardWare_PIP_Z7}"
        sheet['I44'] = f"Z : {object_Inspection.HardWare_PIP_Z8}"

        if object_Inspection.HardWare_Autoload == "Pass":
            sheet['E45'] = "■ Pass          □  Fail"
        elif object_Inspection.HardWare_Autoload == "Fail":
            sheet['E45'] = "□ Pass          ■  Fail"

        if object_Inspection.HardWare_Noise == "Pass":
            sheet['E46'] = "■ Pass          □  Fail"
        elif object_Inspection.HardWare_Noise == "Fail":
            sheet['E46'] = "□ Pass          ■  Fail"

        wb.save('D:\Inst_QC\Report.xlsx')

    def Excel_Inspection3(self):
        # 성적서 3페이지
        object_Inspection = get_object_or_404(Inspection, Instrument_SN=self.kwargs['Instrument_SN'])
        # 수입검사 성적서 엑셀 파일에 데이터 삽입
        wb = openpyxl.load_workbook('D:\Inst_QC\Report.xlsx', data_only=True)
        sheet = wb.active
        sheet = wb['3']

        if object_Inspection.Perfomance_CoverSafety == "Pass":
            sheet['L3'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_CoverSafety == "Fail":
            sheet['L3'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_Barcode == "Pass":
            sheet['L5'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_Barcode == "Fail":
            sheet['L5'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_XYZpositioning == "Pass":
            sheet['L7'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_XYZpositioning == "Fail":
            sheet['L7'] = "□ Pass          ■  Fail"

        sheet.merge_cells('C10:D11')
        sheet.merge_cells('C11:D11')
        sheet['C10'] = object_Inspection.Perfomance_HHS_SN
        sheet["C10"].alignment = Alignment(horizontal="center")

        if object_Inspection.Perfomance_HHS_temperature == "Pass":
            sheet['L8'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_HHS_temperature == "Fail":
            sheet['L8'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_HHS_RPM == "Pass":
            sheet['L11'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_HHS_RPM == "Fail":
            sheet['L11'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_Loading_TipCarrier == "Pass":
            sheet['L12'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_Loading_TipCarrier == "Fail":
            sheet['L12'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_Loading_TubeCarrier == "Pass":
            sheet['L13'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_Loading_TubeCarrier == "Fail":
            sheet['L13'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_Loading_mtpCarrier == "Pass":
            sheet['L14'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_Loading_mtpCarrier == "Fail":
            sheet['L14'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_Loading_4mfxCarrier == "Pass":
            sheet['L15'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_Loading_4mfxCarrier == "Fail":
            sheet['L15'] = "□ Pass          ■  Fail"

        # VFV =====================================================================================
        # P1
        if object_Inspection.Perfomance_VFV_Accuracy300ul1 == "Pass":
            sheet['L16'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy300ul1 == "Fail":
            sheet['L16'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision300ul1 == "Pass":
            sheet['L17'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision300ul1 == "Fail":
            sheet['L17'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Accuracy1000ul1 == "Pass":
            sheet['L18'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy1000ul1 == "Fail":
            sheet['L18'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision1000ul1 == "Pass":
            sheet['L19'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision1000ul1 == "Fail":
            sheet['L19'] = "□ Pass          ■  Fail"

        # P2
        if object_Inspection.Perfomance_VFV_Accuracy300ul2 == "Pass":
            sheet['L20'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy300ul2 == "Fail":
            sheet['L20'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision300ul2 == "Pass":
            sheet['L21'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision300ul2 == "Fail":
            sheet['L21'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Accuracy1000ul2 == "Pass":
            sheet['L22'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy1000ul2 == "Fail":
            sheet['L22'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision1000ul2 == "Pass":
            sheet['L23'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision1000ul2 == "Fail":
            sheet['L23'] = "□ Pass          ■  Fail"

        # P3
        if object_Inspection.Perfomance_VFV_Accuracy300ul3 == "Pass":
            sheet['L24'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy300ul3 == "Fail":
            sheet['L24'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision300ul3 == "Pass":
            sheet['L25'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision300ul3 == "Fail":
            sheet['L25'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Accuracy1000ul3 == "Pass":
            sheet['L26'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy1000ul3 == "Fail":
            sheet['L26'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision1000ul3 == "Pass":
            sheet['L27'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision1000ul3 == "Fail":
            sheet['L27'] = "□ Pass          ■  Fail"

        # P4
        if object_Inspection.Perfomance_VFV_Accuracy300ul4 == "Pass":
            sheet['L28'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy300ul4 == "Fail":
            sheet['L28'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision300ul4 == "Pass":
            sheet['L29'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision300ul4 == "Fail":
            sheet['L29'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Accuracy1000ul4 == "Pass":
            sheet['L30'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy1000ul4 == "Fail":
            sheet['L30'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision1000ul4 == "Pass":
            sheet['L31'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision1000ul4 == "Fail":
            sheet['L31'] = "□ Pass          ■  Fail"

        # P5
        if object_Inspection.Perfomance_VFV_Accuracy300ul5 == "Pass":
            sheet['L32'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy300ul5 == "Fail":
            sheet['L32'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision300ul5 == "Pass":
            sheet['L33'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision300ul5 == "Fail":
            sheet['L33'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Accuracy1000ul5 == "Pass":
            sheet['L34'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy1000ul5 == "Fail":
            sheet['L34'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision1000ul5 == "Pass":
            sheet['L35'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision1000ul5 == "Fail":
            sheet['L35'] = "□ Pass          ■  Fail"

        # P5
        if object_Inspection.Perfomance_VFV_Accuracy300ul5 == "Pass":
            sheet['L32'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy300ul5 == "Fail":
            sheet['L32'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision300ul5 == "Pass":
            sheet['L33'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision300ul5 == "Fail":
            sheet['L33'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Accuracy1000ul5 == "Pass":
            sheet['L34'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy1000ul5 == "Fail":
            sheet['L34'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision1000ul5 == "Pass":
            sheet['L35'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision1000ul5 == "Fail":
            sheet['L35'] = "□ Pass          ■  Fail"

        # P6
        if object_Inspection.Perfomance_VFV_Accuracy300ul6 == "Pass":
            sheet['L36'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy300ul6 == "Fail":
            sheet['L36'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision300ul6 == "Pass":
            sheet['L37'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision300ul6 == "Fail":
            sheet['L37'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Accuracy1000ul6 == "Pass":
            sheet['L38'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy1000ul6 == "Fail":
            sheet['L38'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision1000ul6 == "Pass":
            sheet['L39'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision1000ul6 == "Fail":
            sheet['L39'] = "□ Pass          ■  Fail"

        # P7
        if object_Inspection.Perfomance_VFV_Accuracy300ul7 == "Pass":
            sheet['L40'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy300ul7 == "Fail":
            sheet['L40'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision300ul7 == "Pass":
            sheet['L41'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision300ul7 == "Fail":
            sheet['L41'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Accuracy1000ul7 == "Pass":
            sheet['L42'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy1000ul7 == "Fail":
            sheet['L42'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision1000ul7 == "Pass":
            sheet['L43'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision1000ul7 == "Fail":
            sheet['L43'] = "□ Pass          ■  Fail"

        # P8
        if object_Inspection.Perfomance_VFV_Accuracy300ul8 == "Pass":
            sheet['L44'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy300ul8 == "Fail":
            sheet['L44'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision300ul8 == "Pass":
            sheet['L45'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision300ul8 == "Fail":
            sheet['L45'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Accuracy1000ul8 == "Pass":
            sheet['L46'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Accuracy1000ul8 == "Fail":
            sheet['L46'] = "□ Pass          ■  Fail"

        if object_Inspection.Perfomance_VFV_Precision1000ul8 == "Pass":
            sheet['L47'] = "■ Pass          □  Fail"
        elif object_Inspection.Perfomance_VFV_Precision1000ul8 == "Fail":
            sheet['L47'] = "□ Pass          ■  Fail"

        # Attachment
        if object_Inspection.Attachment_CoverSafety_Report == "Attached":
            sheet['L51'] = "■"
        if object_Inspection.Attachment_Barcode_Report == "Attached":
            sheet['L52'] = "■"
        if object_Inspection.Attachment_Position_Report == "Attached":
            sheet['L53'] = "■"
        if object_Inspection.Attachment_Declaration_HHS == "Attached":
            sheet['L54'] = "■"
        if object_Inspection.Attachment_Declaration == "Attached":
            sheet['L55'] = "■"
        if object_Inspection.Attachment_Measurement_Protocol == "Attached":
            sheet['L56'] = "■"
        if object_Inspection.Attachment_ElectricalSafety_Report == "Attached":
            sheet['L57'] = "■"

        wb.save('D:\Inst_QC\Report.xlsx')

    def Excel_Inspection4(self):
        # 성적서 2페이지
        object_Accesories = get_object_or_404(Accessories, Instrument_SN=self.kwargs['Instrument_SN'])
        object_Inspection = get_object_or_404(Inspection, Instrument_SN=self.kwargs['Instrument_SN'])
        # 수입검사 성적서 엑셀 파일에 데이터 삽입
        wb = openpyxl.load_workbook('D:\Inst_QC\Report.xlsx', data_only=True)
        sheet = wb.active
        sheet = wb['4']

        if object_Accesories.Unpack_Instructions == "checked":
            sheet['K7'] = "■"
        if object_Accesories.Installation_Qualification == "checked":
            sheet['K8'] = "■"
        if object_Accesories.Final_Configuation == "checked":
            sheet['K9'] = "■"
        if object_Accesories.Measurement_Protocol == "checked":
            sheet['K10'] = "■"
        if object_Accesories.EU_Declaration == "checked":
            sheet['K11'] = "■"
        if object_Accesories.DeclarationSTARlet == "checked":
            sheet['K12'] = "■"
        if object_Accesories.Declaration_Quality == "checked":
            sheet['K13'] = "■"
        if object_Accesories.Packing_Checklist == "checked":
            sheet['K14'] = "■"
        if object_Accesories.Maintenance == "checked":
            sheet['K15'] = "■"
        if object_Accesories.Channel_Adjust == "checked":
            sheet['K16'] = "■"
        if object_Accesories.Test_Report == "checked":
            sheet['K17'] = "■"
        if object_Accesories.Loading_Table == "checked":
            sheet['K18'] = "■"
        if object_Accesories.Side_Cover_right == "checked":
            sheet['K19'] = "■"
        if object_Accesories.Side_Cover_left == "checked":
            sheet['K20'] = "■"
        if object_Accesories.Rear_Piexiglas == "checked":
            sheet['K21'] = "■"
        if object_Accesories.Top_Plexiglas == "checked":
            sheet['K22'] = "■"
        if object_Accesories.Left_Plexiglas == "checked":
            sheet['K23'] = "■"
        if object_Accesories.Right_Plexiglas == "checked":
            sheet['K24'] = "■"

        # ACC Kit2
        if object_Accesories.Sensor_4L == "checked":
            sheet['K26'] = "■"
        if object_Accesories.Solid_Waste == "checked":
            sheet['K27'] = "■"
        if object_Accesories.USB_Liquid == "checked":
            sheet['K28'] = "■"
        if object_Accesories.USB_Solid_Liquid == "checked":
            sheet['K29'] = "■"
        if object_Accesories.USB_Venus == "checked":
            sheet['K30'] = "■"
        if object_Accesories.Eppendorf == "checked":
            sheet['K31'] = "■"
        if object_Accesories.Core_Gripper == "checked":
            sheet['K32'] = "■"

            sheet['D33'] = f"Teaching Needle(8ea)  [LOT:       {object_Accesories.Needle_Lot}               ]"

        if object_Accesories.Needle_Check == "checked":
            sheet['K33'] = "■"
        if object_Accesories.Separator == "checked":
            sheet['K34'] = "■"
        if object_Accesories.Tip_Eject == "checked":
            sheet['K35'] = "■"
        if object_Accesories.Power_Cord7 == "checked":
            sheet['K36'] = "■"
        if object_Accesories.Power_CordP == "checked":
            sheet['K37'] = "■"
        if object_Accesories.Cable_USB == "checked":
            sheet['K38'] = "■"
        if object_Accesories.Fuse == "checked":
            sheet['K39'] = "■"
        if object_Accesories.Screw == "checked":
            sheet['K40'] = "■"
        if object_Accesories.Fitting == "checked":
            sheet['K41'] = "■"
        if object_Accesories.Liq_Waste == "checked":
            sheet['K42'] = "■"

        # ACC Kit3
        if object_Accesories.XArm == "checked":
            sheet['K44'] = "■"
        if object_Accesories.Bar_Left == "checked":
            sheet['K45'] = "■"
        if object_Accesories.Bar_Top == "checked":
            sheet['K46'] = "■"

        if object_Accesories.Bag == "checked":
            sheet['K48'] = "■"
        if object_Accesories.Tube_32 == "checked":
            sheet['K49'] = "■"
        if object_Accesories.TipRack_5 == "checked":
            sheet['K50'] = "■"
        if object_Accesories.MFX4 == "checked":
            sheet['K51'] = "■"
        if object_Accesories.SEEG == "checked":
            sheet['K52'] = "■"
        if object_Accesories.FLHX == "checked":
            sheet['K53'] = "■"
        if object_Accesories.SOHX == "checked":
            sheet['K54'] = "■"
        if object_Accesories.HHS_Plate == "checked":
            sheet['K55'] = "■"
        if object_Accesories.MTP == "checked":
            sheet['K56'] = "■"
        if object_Accesories.Verify_Kit == "checked":
            sheet['K57'] = "■"
        if object_Accesories.STD == "checked":
            sheet['K58'] = "■"
        if object_Accesories.High == "checked":
            sheet['K59'] = "■"

        sheet['D60'] = f"Hamiltion Heater & Shaker (HHS)   [SN :         {object_Accesories.HHS_SN}               ]"

        if object_Accesories.HHS_Check == "checked":
            sheet['K60'] = "■"
        if object_Accesories.HHS_Manual == "checked":
            sheet['K61'] = "■"
        if object_Accesories.Stop_Barcode == "checked":
            sheet['K62'] = "■"

        sheet['J4'] = object_Inspection.Computer_SN

        wb.save('D:\Inst_QC\Report.xlsx')